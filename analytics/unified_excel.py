"""Update the owner's real "Análisis unificado" workbook (US30).

The owner keeps a master workbook whose ``Productos vendidos`` sheet is a units
matrix with a **two-row header**: one row of years and, below it, one row of
month names (``Enero``..``Diciembre``), repeated per year. Rows are
``Grupo / Clave / Producto`` and each cell is the raw POS ``CANTIDAD`` for that
month (grams for cuts sold by weight, units for fixed portions) -- exactly what
Mesa already stores, so no conversion is needed.

:func:`update_productos_vendidos` fills the column for one ``(year, month)`` with
Mesa's per-product units for that period. Products are matched to existing rows
by NAME (the same fusion logic as the importer, US22) because POS claves are
unreliable and the historical spreadsheet names drift from the POS names; the
historical clave is never overwritten. Unmatched products are appended as new
rows. The result is written to a COPY -- the original is never modified.

Scope (agreed): only the ``Productos vendidos`` sheet. The ``Datos totales``
sheet (N/O/S/T from the PDF footer) is a separate story.
"""
from __future__ import annotations

import datetime
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from catalog.identity import names_match, normalize_name

from .exports import MONTHS_ES, PRODUCTOS_VENDIDOS_SHEET, _matrix_data

_GROUP_COL = 1
_CLAVE_COL = 2
_NAME_COL = 3
_FIRST_MONTH_COL = 4

# --- "Datos totales" sheet (US32) -----------------------------------------
_DATOS_TOTALES_SHEET = "Datos totales"  # matched ignoring a trailing space
_DT_YEAR = 1
_DT_DATE = 2
_DT_DAY = 3
_DT_MONTH = 4
_DT_WEEKDAY = 5
_DT_VENTA_BAR = 14      # N
_DT_COSTO_BAR = 15      # O
_DT_VENTA_COCINA = 19   # S
_DT_COSTO_COCINA = 20   # T
_WEEKDAYS_ES = (
    "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"
)


class UnifiedUpdateError(Exception):
    """The workbook cannot be updated (missing file, wrong layout, ...)."""


def _norm_clave(value) -> str:
    """Compare claves ignoring leading zeros (``"03028"`` == ``3028``)."""
    if value is None:
        return ""
    text = str(value).strip()
    return text.lstrip("0") or "0"


def _norm_month(value) -> str:
    return str(value).strip().capitalize() if value is not None else ""


def _locate_header(ws) -> tuple[int, int]:
    """Return ``(year_row, month_row)`` for the two-row header.

    The month row is the one whose first cell reads ``Grupo``; the year row sits
    directly above it. Raises when the layout is not recognized.
    """
    for row in range(1, 11):
        if str(ws.cell(row=row, column=_GROUP_COL).value or "").strip().lower() == "grupo":
            if row < 2:
                raise UnifiedUpdateError("No year row above the 'Grupo' header row.")
            return row - 1, row
    raise UnifiedUpdateError(
        f'Could not find the "Grupo/Clave/Producto" header in '
        f'"{PRODUCTOS_VENDIDOS_SHEET}".'
    )


def _find_month_column(ws, year_row: int, month_row: int, year: int, month: int) -> int:
    """Column index for ``(year, month)``, appending a new one if absent."""
    target_name = MONTHS_ES[month - 1]
    for col in range(_FIRST_MONTH_COL, ws.max_column + 1):
        cell_year = ws.cell(row=year_row, column=col).value
        cell_month = _norm_month(ws.cell(row=month_row, column=col).value)
        if str(cell_year).strip() == str(year) and cell_month == target_name:
            return col
    # Not in the grid: append a fresh two-row-header column at the end.
    col = ws.max_column + 1
    ws.cell(row=year_row, column=col, value=year)
    ws.cell(row=month_row, column=col, value=target_name)
    return col


def _existing_rows(ws, data_start: int) -> list[dict]:
    rows = []
    for r in range(data_start, ws.max_row + 1):
        name = ws.cell(row=r, column=_NAME_COL).value
        if name is None or str(name).strip() == "":
            continue
        rows.append(
            {
                "row": r,
                "norm": normalize_name(str(name)),
                "clave": _norm_clave(ws.cell(row=r, column=_CLAVE_COL).value),
            }
        )
    return rows


def _find_row(existing: list[dict], norm: str, clave: str) -> dict | None:
    """Match a product against existing rows: exact-normalized, then fusion."""
    for row in existing:
        if row["norm"] == norm:
            return row
    for row in existing:
        if names_match(norm, row["norm"], same_clave=(clave != "" and clave == row["clave"])):
            return row
    return None


def _detect_warnings(path: Path, workbook) -> list[str]:
    warnings = []
    lock = path.parent / f"~${path.name}"
    if lock.exists():
        warnings.append(
            f"'{path.name}' looks open in Excel (lock file {lock.name}); close it "
            "so the copy reflects the latest saved state."
        )
    if any(getattr(sheet, "_charts", None) for sheet in workbook.worksheets):
        warnings.append("workbook has charts; openpyxl does not preserve them.")
    if path.suffix.lower() == ".xlsm":
        warnings.append("workbook may contain macros; openpyxl does not preserve them.")
    return warnings


def update_productos_vendidos(
    path, restaurant, year: int, month: int
) -> dict:
    """Fill the ``(year, month)`` column of the units matrix in a copy of ``path``.

    Returns a summary with ``copy`` (written path), ``column`` (label like
    ``"AP / Marzo 2025"``), ``matched``/``appended`` counts, ``appended_names``
    (for review of possible name-drift mismatches) and ``warnings``.
    """
    if not 1 <= month <= 12:
        raise UnifiedUpdateError(f"Month out of range: {month}")
    path = Path(path)
    if not path.exists():
        raise UnifiedUpdateError(f"File not found: {path}")

    workbook = load_workbook(path)  # keep formulas (data_only=False)
    if PRODUCTOS_VENDIDOS_SHEET not in workbook.sheetnames:
        raise UnifiedUpdateError(
            f'Sheet "{PRODUCTOS_VENDIDOS_SHEET}" not found in {path.name}.'
        )
    ws = workbook[PRODUCTOS_VENDIDOS_SHEET]
    warnings = _detect_warnings(path, workbook)

    year_row, month_row = _locate_header(ws)
    data_start = month_row + 1
    col = _find_month_column(ws, year_row, month_row, year, month)

    _, products = _matrix_data(restaurant)
    period = (year, month)
    to_write = [p for p in products.values() if p["quantities"].get(period)]

    existing = _existing_rows(ws, data_start)
    matched = 0
    appended_names: list[str] = []
    append_at = ws.max_row + 1
    for product in sorted(to_write, key=lambda p: (p["grupo"] or "", p["name"])):
        norm = normalize_name(product["name"])
        clave = _norm_clave(product["clave"])
        match = _find_row(existing, norm, clave)
        if match is None:
            target = append_at
            append_at += 1
            ws.cell(row=target, column=_GROUP_COL, value=product["grupo"])
            ws.cell(row=target, column=_CLAVE_COL, value=product["clave"])
            ws.cell(row=target, column=_NAME_COL, value=product["name"])
            existing.append({"row": target, "norm": norm, "clave": clave})
            appended_names.append(product["name"])
        else:
            target = match["row"]
            ws.cell(row=target, column=_GROUP_COL, value=product["grupo"])  # refresh group
            matched += 1
        ws.cell(row=target, column=col, value=product["quantities"][period])

    copy = path.with_name(f"{path.stem} (Mesa {MONTHS_ES[month - 1]} {year}){path.suffix}")
    workbook.save(copy)
    return {
        "copy": copy,
        "column": f"{get_column_letter(col)} / {MONTHS_ES[month - 1]} {year}",
        "matched": matched,
        "appended": len(appended_names),
        "appended_names": appended_names,
        "warnings": warnings,
    }


def _find_datos_sheet(workbook):
    for name in workbook.sheetnames:
        if name.strip().lower() == _DATOS_TOTALES_SHEET.lower():
            return workbook[name]
    raise UnifiedUpdateError(f'Sheet "{_DATOS_TOTALES_SHEET}" not found.')


def _index_dates(ws) -> dict[datetime.date, int]:
    """Map each existing dated row (col B) to its row number."""
    index = {}
    for r in range(2, ws.max_row + 1):
        value = ws.cell(row=r, column=_DT_DATE).value
        if isinstance(value, datetime.datetime):
            index[value.date()] = r
        elif isinstance(value, datetime.date):
            index[value] = r
    return index


def _find_formula_template(ws) -> int | None:
    """A data row whose formula cells (J=N+S, ...) can be copied to new rows."""
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=10)  # column J = "Venta Total" (a formula)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            return r
    return None


def _copy_row_formulas(ws, template: int, target: int) -> None:
    """Replicate the template row's per-row formulas onto ``target``.

    Formulas here are same-row references (``=N2+S2``); openpyxl does not
    translate them on copy, so rewrite the template's row number to the target's.
    """
    pattern = re.compile(r"([A-Z]{1,3})" + str(template) + r"(?![0-9])")
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=template, column=col).value
        if isinstance(value, str) and value.startswith("="):
            rewritten = pattern.sub(lambda m: f"{m.group(1)}{target}", value)
            ws.cell(row=target, column=col, value=rewritten)


def _write_date_columns(ws, row: int, day: datetime.date) -> None:
    ws.cell(row=row, column=_DT_YEAR, value=day.year)
    # Excel has no concept of a timezone: openpyxl serialises a datetime to a plain
    # serial number, and a tz-aware one would be rejected. Naive is correct here.
    ws.cell(
        row=row,
        column=_DT_DATE,
        value=datetime.datetime(day.year, day.month, day.day),  # noqa: DTZ001
    )
    ws.cell(row=row, column=_DT_DAY, value=day.day)
    ws.cell(row=row, column=_DT_MONTH, value=MONTHS_ES[day.month - 1])
    ws.cell(row=row, column=_DT_WEEKDAY, value=_WEEKDAYS_ES[day.weekday()])


def update_datos_totales(path, totals_by_date: dict) -> dict:
    """Write Bar/Kitchen totals (N/O/S/T) per day into a copy of ``path``.

    ``totals_by_date`` maps a ``datetime.date`` to a
    :class:`sales.importers.pdf_daily.DailyTotals`. Existing rows (matched by the
    date in column B) are filled in place; days with no row are appended at the
    bottom with the date columns and the per-row formulas replicated, and are
    reported in ``appended`` (openpyxl cannot safely insert mid-sheet without
    corrupting the shifted rows' formulas). Writes a copy; returns a summary.
    """
    path = Path(path)
    if not path.exists():
        raise UnifiedUpdateError(f"File not found: {path}")

    workbook = load_workbook(path)
    ws = _find_datos_sheet(workbook)
    warnings = _detect_warnings(path, workbook)

    existing = _index_dates(ws)
    template = _find_formula_template(ws)
    append_at = ws.max_row + 1
    filled = 0
    appended: list[datetime.date] = []

    for day in sorted(totals_by_date):
        totals = totals_by_date[day]
        row = existing.get(day)
        if row is None:
            row = append_at
            append_at += 1
            _write_date_columns(ws, row, day)
            if template is not None:
                _copy_row_formulas(ws, template, row)
            appended.append(day)
        else:
            filled += 1
        ws.cell(row=row, column=_DT_VENTA_BAR, value=float(totals.venta_bar))
        ws.cell(row=row, column=_DT_COSTO_BAR, value=float(totals.costo_bar))
        ws.cell(row=row, column=_DT_VENTA_COCINA, value=float(totals.venta_cocina))
        ws.cell(row=row, column=_DT_COSTO_COCINA, value=float(totals.costo_cocina))

    if appended and template is None:
        warnings.append(
            "No formula template row found; appended rows carry N/O/S/T only "
            "(the derived columns J/K/L... will be empty)."
        )

    copy = path.with_name(f"{path.stem} (Mesa Datos totales){path.suffix}")
    workbook.save(copy)
    return {
        "copy": copy,
        "filled": filled,
        "appended": [d.isoformat() for d in appended],
        "warnings": warnings,
    }
