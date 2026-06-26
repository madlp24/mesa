"""Update an existing "Productos vendidos" workbook in place (US23).

Takes the owner's real ``.xlsx`` (matrix sheet: cols Grupo/Clave/Producto, then
one column per month x year) and writes the months it does not yet have, matching
existing rows by NAME with the same fusion logic as the importer (US22). Truly
new products are appended as new rows; historical codes are never overwritten.
The result is written to a COPY, never the original.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from catalog.identity import names_match, normalize_name

from .exports import PRODUCTOS_VENDIDOS_SHEET, MONTHS_ES, _matrix_data, month_label

_GROUP_COL = 1
_CLAVE_COL = 2
_NAME_COL = 3
_FIRST_MONTH_COL = 4


class ExcelUpdateError(Exception):
    """The workbook cannot be updated (missing, wrong sheet, ...)."""


def parse_month_label(label) -> tuple[int, int] | None:
    """Parse a ``"Enero 2026"`` header into ``(year, month)``, else ``None``."""
    if not label:
        return None
    parts = str(label).strip().split()
    if len(parts) != 2 or not parts[1].isdigit():
        return None
    month_name = parts[0].capitalize()
    if month_name not in MONTHS_ES:
        return None
    return int(parts[1]), MONTHS_ES.index(month_name) + 1


def _find_row(existing_rows: list[dict], norm: str, clave: str) -> dict | None:
    """Match a product against existing rows: exact-normalized, then fusion."""
    for row in existing_rows:
        if row["norm"] == norm:
            return row
    for row in existing_rows:
        if names_match(norm, row["norm"], same_clave=(clave == row["clave"])):
            return row
    return None


def _detect_warnings(path: Path, workbook) -> list[str]:
    warnings = []
    lock = path.parent / f"~${path.name}"
    if lock.exists():
        warnings.append(
            f"'{path.name}' looks open in Excel (lock file {lock.name}); "
            "close it so the copy reflects the latest saved state."
        )
    if any(getattr(sheet, "_charts", None) for sheet in workbook.worksheets):
        warnings.append("workbook has charts; openpyxl does not preserve them.")
    if path.suffix.lower() == ".xlsm":
        warnings.append("workbook may contain macros; openpyxl does not preserve them.")
    return warnings


def update_productos_vendidos(path: Path, restaurant) -> dict:
    """Write the missing months into a copy of ``path``; return a summary.

    Uses the given restaurant's data. The summary has ``copy`` (the written
    path), ``months_added`` (labels), ``matched`` / ``appended`` row counts, and
    ``warnings``.
    """
    path = Path(path)
    if not path.exists():
        raise ExcelUpdateError(f"File not found: {path}")

    workbook = load_workbook(path)
    if PRODUCTOS_VENDIDOS_SHEET not in workbook.sheetnames:
        raise ExcelUpdateError(
            f'Sheet "{PRODUCTOS_VENDIDOS_SHEET}" not found in {path.name}'
        )
    ws = workbook[PRODUCTOS_VENDIDOS_SHEET]
    warnings = _detect_warnings(path, workbook)

    header = [cell.value for cell in ws[1]]
    existing_months = {}
    for offset, label in enumerate(header[_FIRST_MONTH_COL - 1 :]):
        period = parse_month_label(label)
        if period:
            existing_months[period] = _FIRST_MONTH_COL + offset

    months, products = _matrix_data(restaurant)
    new_months = [m for m in months if m not in existing_months]

    # Append a column per new month.
    month_cols = {}
    next_col = ws.max_column + 1
    for period in new_months:
        ws.cell(row=1, column=next_col, value=month_label(*period))
        month_cols[period] = next_col
        next_col += 1

    # Index existing product rows by normalized name.
    existing_rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=_NAME_COL).value
        if name is None:
            continue
        clave = ws.cell(row=r, column=_CLAVE_COL).value
        existing_rows.append(
            {
                "row": r,
                "norm": normalize_name(str(name)),
                "clave": "" if clave is None else str(clave),
            }
        )

    matched = 0
    appended = 0
    append_at = ws.max_row + 1
    for product in products.values():
        norm = normalize_name(product["name"])
        clave = str(product["clave"])
        match = _find_row(existing_rows, norm, clave)
        if match is None:
            target = append_at
            append_at += 1
            ws.cell(row=target, column=_GROUP_COL, value=product["grupo"])
            ws.cell(row=target, column=_CLAVE_COL, value=product["clave"])
            ws.cell(row=target, column=_NAME_COL, value=product["name"])
            existing_rows.append({"row": target, "norm": norm, "clave": clave})
            appended += 1
        else:
            target = match["row"]
            # Refresh the group to the latest report; never touch the code.
            ws.cell(row=target, column=_GROUP_COL, value=product["grupo"])
            matched += 1
        for period, col in month_cols.items():
            quantity = product["quantities"].get(period)
            if quantity is not None:
                ws.cell(row=target, column=col, value=quantity)

    copy = path.with_name(f"{path.stem} (actualizado){path.suffix}")
    workbook.save(copy)
    return {
        "copy": copy,
        "months_added": [month_label(*m) for m in new_months],
        "matched": matched,
        "appended": appended,
        "warnings": warnings,
    }
