"""Tests for the 'Datos totales' N/O/S/T updater (US32)."""
import datetime
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook

from analytics.unified_excel import UnifiedUpdateError, update_datos_totales
from sales.importers.pdf_daily import DailyTotals


def _totals(d):
    return DailyTotals(
        date=d,
        venta_bar=Decimal("100"),
        costo_bar=Decimal("40"),
        venta_cocina=Decimal("300"),
        costo_cocina=Decimal("120"),
    )


def _master(path):
    """A minimal 'Datos totales' sheet with two dated rows and per-row formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos totales "  # trailing space, like the real file
    ws.cell(row=1, column=1, value="Año")
    ws.cell(row=1, column=2, value="Fecha")
    # Two existing daily rows with the same per-row formulas as the real sheet.
    for r, day in ((2, datetime.datetime(2025, 3, 5)), (3, datetime.datetime(2025, 3, 10))):
        ws.cell(row=r, column=1, value=day.year)
        ws.cell(row=r, column=2, value=day)
        ws.cell(row=r, column=10, value=f"=N{r}+S{r}")  # J Venta Total
        ws.cell(row=r, column=11, value=f"=O{r}+T{r}")  # K Costo Total
        ws.cell(row=r, column=12, value=f"=J{r}-K{r}")  # L Utilidad
    wb.save(path)


@pytest.fixture
def master(tmp_path):
    path = tmp_path / "Analisis.xlsx"
    _master(path)
    return path


def test_fills_existing_date_row_in_place(master):
    summary = update_datos_totales(
        master, {datetime.date(2025, 3, 5): _totals("2025-03-05")}
    )

    assert summary["filled"] == 1
    assert summary["appended"] == []
    ws = load_workbook(summary["copy"])["Datos totales "]
    assert ws.cell(row=2, column=14).value == 100  # N Venta Bar
    assert ws.cell(row=2, column=15).value == 40   # O Costo Bar
    assert ws.cell(row=2, column=19).value == 300  # S Venta Cocina
    assert ws.cell(row=2, column=20).value == 120  # T Costo Cocina


def test_appends_new_date_with_rewired_formulas(master):
    summary = update_datos_totales(
        master, {datetime.date(2025, 3, 20): _totals("2025-03-20")}
    )

    assert summary["filled"] == 0
    assert summary["appended"] == ["2025-03-20"]
    ws = load_workbook(summary["copy"])["Datos totales "]
    new_rows = [
        r for r in range(2, ws.max_row + 1)
        if isinstance(ws.cell(row=r, column=2).value, datetime.datetime)
        and ws.cell(row=r, column=2).value.date() == datetime.date(2025, 3, 20)
    ]
    assert len(new_rows) == 1
    r = new_rows[0]
    assert ws.cell(row=r, column=1).value == 2025          # year
    assert ws.cell(row=r, column=4).value == "Marzo"       # month name ES
    assert ws.cell(row=r, column=5).value == "Jueves"      # 2025-03-20 is Thursday
    assert ws.cell(row=r, column=14).value == 100          # N
    # Formula was rewired from the template row to this row.
    assert ws.cell(row=r, column=10).value == f"=N{r}+S{r}"


def test_both_fill_and_append(master):
    summary = update_datos_totales(
        master,
        {
            datetime.date(2025, 3, 10): _totals("2025-03-10"),  # existing
            datetime.date(2025, 3, 25): _totals("2025-03-25"),  # new
        },
    )
    assert summary["filled"] == 1
    assert summary["appended"] == ["2025-03-25"]


def test_original_not_modified(master):
    update_datos_totales(master, {datetime.date(2025, 3, 5): _totals("2025-03-05")})
    original = load_workbook(master)["Datos totales "]
    assert original.cell(row=2, column=14).value is None


def test_missing_sheet_raises(tmp_path):
    path = tmp_path / "wrong.xlsx"
    wb = Workbook()
    wb.active.title = "Otra"
    wb.save(path)
    with pytest.raises(UnifiedUpdateError, match="Datos totales"):
        update_datos_totales(path, {datetime.date(2025, 3, 5): _totals("2025-03-05")})
