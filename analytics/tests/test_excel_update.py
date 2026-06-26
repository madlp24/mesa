"""Tests for updating an existing Productos-Vendidos workbook (US23)."""
from datetime import datetime, timezone
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook

from analytics.excel_update import (
    ExcelUpdateError,
    parse_month_label,
    update_productos_vendidos,
)
from catalog.models import Category, Product
from sales.models import Sale, SaleItem


def _sample_workbook(path):
    """A workbook with one existing month (Enero 2026) and two product rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos vendidos"
    ws.append(["Grupo", "Clave", "Producto", "Enero 2026"])
    ws.append(["COCTELES", "8100", "Negroni", 5])  # later "Negroni Tanqueray"
    ws.append(["ENTRADAS", "03028", "Arepa de Choclo", 10])
    wb.save(path)


def _sell(product, ext, year, month, day, qty):
    sale = Sale.objects.create(
        external_id=ext,
        occurred_at=datetime(year, month, day, 12, tzinfo=timezone.utc),
        total=product.sale_price * qty,
    )
    SaleItem.objects.create(
        sale=sale, product=product, quantity=qty,
        unit_price=product.sale_price, unit_cost=product.cost_price,
    )


@pytest.fixture
def seeded_db(db):
    cocteles = Category.objects.create(name="COCTELES")
    entradas = Category.objects.create(name="ENTRADAS")
    nuevos = Category.objects.create(name="NUEVOS")
    negroni = Product.objects.create(
        name="Negroni Tanqueray", sku="8100", category=cocteles,
        cost_price=Decimal("7"), sale_price=Decimal("22"),
    )
    arepa = Product.objects.create(
        name="Arepa de Choclo", sku="03028", category=entradas,
        cost_price=Decimal("3"), sale_price=Decimal("10"),
    )
    limonada = Product.objects.create(
        name="Limonada", sku="5000", category=nuevos,
        cost_price=Decimal("1"), sale_price=Decimal("5"),
    )
    # New month (February) for all three; plus a January sale to prove the
    # existing month column is not overwritten.
    _sell(negroni, "n-feb", 2026, 2, 3, 7)
    _sell(arepa, "a-feb", 2026, 2, 4, 9)
    _sell(limonada, "l-feb", 2026, 2, 5, 12)
    _sell(negroni, "n-jan", 2026, 1, 9, 99)
    return {"negroni": negroni, "arepa": arepa, "limonada": limonada}


def test_parse_month_label():
    assert parse_month_label("Enero 2026") == (2026, 1)
    assert parse_month_label("DICIEMBRE 2025") == (2025, 12)
    assert parse_month_label("Grupo") is None
    assert parse_month_label(None) is None


@pytest.mark.django_db
def test_update_writes_copy_not_original(tmp_path, seeded_db):
    original = tmp_path / "ventas.xlsx"
    _sample_workbook(original)

    summary = update_productos_vendidos(original)

    assert summary["copy"] == tmp_path / "ventas (actualizado).xlsx"
    assert summary["copy"].exists()
    # Original is untouched: still only the January column.
    ws = load_workbook(original)["Productos vendidos"]
    assert [c.value for c in ws[1]] == ["Grupo", "Clave", "Producto", "Enero 2026"]


@pytest.mark.django_db
def test_new_month_appended_and_existing_preserved(tmp_path, seeded_db):
    path = tmp_path / "ventas.xlsx"
    _sample_workbook(path)

    summary = update_productos_vendidos(path)
    assert summary["months_added"] == ["Febrero 2026"]

    ws = load_workbook(summary["copy"])["Productos vendidos"]
    header = [c.value for c in ws[1]]
    assert header == ["Grupo", "Clave", "Producto", "Enero 2026", "Febrero 2026"]

    rows = {r[2]: r for r in ws.iter_rows(min_row=2, values_only=True)}
    # Negroni: matched the "Negroni" row by prefix+clave; February written,
    # January NOT overwritten (stays 5, not the 99 from the Jan sale).
    assert rows["Negroni"][3] == 5
    assert rows["Negroni"][4] == 7
    # Code preserved on the matched row.
    assert rows["Negroni"][1] == "8100"
    assert rows["Arepa de Choclo"][4] == 9


@pytest.mark.django_db
def test_new_product_appended_as_new_row(tmp_path, seeded_db):
    path = tmp_path / "ventas.xlsx"
    _sample_workbook(path)

    summary = update_productos_vendidos(path)
    assert summary["appended"] == 1
    assert summary["matched"] == 2

    ws = load_workbook(summary["copy"])["Productos vendidos"]
    rows = {r[2]: r for r in ws.iter_rows(min_row=2, values_only=True)}
    assert "Limonada" in rows
    assert rows["Limonada"][1] == "5000"  # new code on a new row
    assert rows["Limonada"][4] == 12  # February quantity


@pytest.mark.django_db
def test_warns_when_workbook_locked(tmp_path, seeded_db):
    path = tmp_path / "ventas.xlsx"
    _sample_workbook(path)
    (tmp_path / "~$ventas.xlsx").write_text("lock")

    summary = update_productos_vendidos(path)
    assert any("open in Excel" in w for w in summary["warnings"])


@pytest.mark.django_db
def test_missing_sheet_raises(tmp_path, seeded_db):
    path = tmp_path / "wrong.xlsx"
    wb = Workbook()
    wb.active.title = "Otra hoja"
    wb.save(path)

    with pytest.raises(ExcelUpdateError, match="not found"):
        update_productos_vendidos(path)
