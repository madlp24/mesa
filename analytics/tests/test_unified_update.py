"""Tests for the unified-analysis 'Productos vendidos' updater (US30)."""
from datetime import UTC, datetime
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook

from analytics.unified_excel import UnifiedUpdateError, update_productos_vendidos
from catalog.models import Category, Product
from sales.models import Sale, SaleItem


def _master_workbook(path):
    """A minimal two-row-header 'Productos vendidos' workbook like the owner's."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos vendidos"
    # Row 1 blank; row 2 = years; row 3 = Grupo/Clave/Producto + month names.
    ws.cell(row=2, column=4, value=2025)  # D2 (Enero 2025)
    ws.cell(row=2, column=5, value=2025)  # E2 (Febrero 2025)
    ws.cell(row=2, column=6, value=2025)  # F2 (Marzo 2025)
    for col, label in enumerate(
        ["Grupo", "Clave", "Producto", "Enero", "Febrero", "Marzo"], start=1
    ):
        ws.cell(row=3, column=col, value=label)
    # Existing rows: one matches by name (drifted case/accents), one is unrelated.
    ws.cell(row=4, column=1, value="ENTRADAS")
    ws.cell(row=4, column=2, value=1001)
    ws.cell(row=4, column=3, value="Arepa de Choclo")
    ws.cell(row=5, column=1, value="POSTRES")
    ws.cell(row=5, column=2, value=2001)
    ws.cell(row=5, column=3, value="Brownie de Sarten")
    wb.save(path)


@pytest.fixture
def master(tmp_path):
    path = tmp_path / "Analisis.xlsx"
    _master_workbook(path)
    return path


def _sell(restaurant, category, name, sku, qty, day=5):
    product = Product.objects.create(
        restaurant=restaurant, name=name, sku=sku, category=category,
        cost_price=Decimal("1"), sale_price=Decimal("5"),
    )
    sale = Sale.objects.create(
        restaurant=restaurant, external_id=f"2025-03-{day:02d}:{sku}",
        occurred_at=datetime(2025, 3, day, 12, tzinfo=UTC), total=Decimal("5"),
    )
    SaleItem.objects.create(
        sale=sale, product=product, quantity=qty,
        unit_price=Decimal("5"), unit_cost=Decimal("1"),
    )
    return product


@pytest.mark.django_db
def test_fills_matched_row_and_appends_new(master, restaurant):
    entradas = Category.objects.create(restaurant=restaurant, name="ENTRADAS")
    # "AREPA DE CHOCLO*GR" normalizes to "AREPA DE CHOCLO" -> matches existing row.
    _sell(restaurant, entradas, "AREPA DE CHOCLO*GR", "03028", 68)
    _sell(restaurant, entradas, "TACOS DE BIRRIA *4 UND", "01024", 12)

    summary = update_productos_vendidos(master, restaurant, 2025, 3)

    assert summary["matched"] == 1
    assert summary["appended"] == 1
    assert "TACOS DE BIRRIA *4 UND" in summary["appended_names"]
    assert summary["column"].startswith("F ")  # Marzo 2025 is column F

    # Original untouched; the copy carries the values.
    assert summary["copy"] != master
    ws = load_workbook(summary["copy"])["Productos vendidos"]
    assert ws.cell(row=4, column=6).value == 68  # matched Arepa -> Marzo (F)
    assert ws.cell(row=4, column=2).value == 1001  # historical clave preserved
    # Appended row at the bottom with its own clave and the units.
    appended = [r for r in range(4, ws.max_row + 1)
                if ws.cell(row=r, column=3).value == "TACOS DE BIRRIA *4 UND"]
    assert appended and ws.cell(row=appended[0], column=6).value == 12


@pytest.mark.django_db
def test_appends_month_column_when_absent(master, restaurant):
    cat = Category.objects.create(restaurant=restaurant, name="ENTRADAS")
    _sell(restaurant, cat, "AREPA DE CHOCLO", "03028", 5, day=6)

    # April 2025 is not in the grid -> a new column is appended.
    summary = update_productos_vendidos(master, restaurant, 2025, 4)

    ws = load_workbook(summary["copy"])["Productos vendidos"]
    assert summary["column"].endswith("Abril 2025")
    col = ws.max_column
    assert ws.cell(row=2, column=col).value == 2025
    assert ws.cell(row=3, column=col).value == "Abril"


@pytest.mark.django_db
def test_missing_sheet_raises(tmp_path, restaurant):
    path = tmp_path / "wrong.xlsx"
    wb = Workbook()
    wb.active.title = "Otra hoja"
    wb.save(path)
    with pytest.raises(UnifiedUpdateError, match="not found"):
        update_productos_vendidos(path, restaurant, 2025, 3)


@pytest.mark.django_db
def test_bad_month_raises(master, restaurant):
    with pytest.raises(UnifiedUpdateError, match="Month out of range"):
        update_productos_vendidos(master, restaurant, 2025, 13)


@pytest.mark.django_db
def test_original_file_is_not_modified(master, restaurant):
    cat = Category.objects.create(restaurant=restaurant, name="ENTRADAS")
    _sell(restaurant, cat, "AREPA DE CHOCLO", "03028", 99)

    update_productos_vendidos(master, restaurant, 2025, 3)

    original = load_workbook(master)["Productos vendidos"]
    assert original.cell(row=4, column=6).value is None  # Marzo cell still empty
