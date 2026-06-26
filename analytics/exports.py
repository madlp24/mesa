"""Excel exports built from the database with openpyxl.

Two workbooks:

* :func:`build_productos_vendidos_workbook` -- the "Productos vendidos" matrix
  (rows = Grupo / Clave / Producto, one column per month x year, cell = units
  sold), mirroring the owner's unified-analysis workbook.
* :func:`build_analysis_workbook` -- an analysis report (per product and per
  category: quantity, revenue, cost, margin $ and margin %) plus top-N rankings,
  reusing :mod:`analytics.services`.
"""
from __future__ import annotations

import datetime
from decimal import Decimal

from django.db.models import Sum
from django.db.models.functions import TruncMonth
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from sales.models import SaleItem

from .services import category_report, product_report

MONTHS_ES = (
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
)
PRODUCTOS_VENDIDOS_SHEET = "Productos vendidos"
_HEADER_FONT = Font(bold=True)


def month_label(year: int, month: int) -> str:
    """A column label like ``"Enero 2026"``."""
    return f"{MONTHS_ES[month - 1]} {year}"


def _write_header(ws: Worksheet, headers: list) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT


def _matrix_data(restaurant) -> tuple[list[tuple[int, int]], dict]:
    """Aggregate units sold per product per month for one restaurant.

    Returns ``(months, products)`` where ``months`` is a sorted list of
    ``(year, month)`` and ``products`` maps a product sku to its group, name and
    per-month quantities.
    """
    rows = (
        SaleItem.objects.filter(sale__restaurant=restaurant)
        .annotate(month=TruncMonth("sale__occurred_at"))
        .values(
            "product__sku", "product__name", "product__category__name", "month"
        )
        .annotate(quantity=Sum("quantity"))
    )
    months: set[tuple[int, int]] = set()
    products: dict[str, dict] = {}
    for row in rows:
        period = (row["month"].year, row["month"].month)
        months.add(period)
        sku = row["product__sku"]
        product = products.setdefault(
            sku,
            {
                "grupo": row["product__category__name"],
                "clave": sku,
                "name": row["product__name"],
                "quantities": {},
            },
        )
        product["quantities"][period] = row["quantity"] or 0
    return sorted(months), products


def build_productos_vendidos_workbook(restaurant) -> Workbook:
    """Build the Productos-Vendidos matrix workbook for one restaurant."""
    months, products = _matrix_data(restaurant)
    workbook = Workbook()
    ws = workbook.active
    ws.title = PRODUCTOS_VENDIDOS_SHEET

    headers = ["Grupo", "Clave", "Producto"]
    headers += [month_label(year, month) for year, month in months]
    _write_header(ws, headers)

    ordered = sorted(products.values(), key=lambda p: (p["grupo"] or "", p["name"]))
    for product in ordered:
        row = [product["grupo"], product["clave"], product["name"]]
        row += [product["quantities"].get(period, 0) for period in months]
        ws.append(row)

    _autosize(ws)
    return workbook


def _money(value) -> float:
    return float(value or Decimal("0"))


def _pct(value) -> float:
    return round(float(value or Decimal("0")), 1)


def build_analysis_workbook(
    restaurant,
    start: datetime.date | None = None,
    end: datetime.date | None = None,
    top_n: int = 10,
) -> Workbook:
    """Build the analysis report workbook for one restaurant and date window."""
    products = product_report(restaurant, start, end)
    categories = category_report(restaurant, start, end)

    workbook = Workbook()
    _fill_product_sheet(workbook.active, products)
    _fill_category_sheet(workbook.create_sheet("Por categoría"), categories)
    _fill_rankings_sheet(workbook.create_sheet("Rankings"), products, top_n)
    return workbook


def _fill_product_sheet(ws: Worksheet, products: list[dict]) -> None:
    ws.title = "Por producto"
    _write_header(
        ws,
        ["Producto", "Categoría", "Cantidad", "Ingreso", "Costo", "Margen $", "Margen %"],
    )
    for row in products:
        ws.append(
            [
                row["name"],
                row["category"],
                row["quantity"],
                _money(row["revenue"]),
                _money(row["cost"]),
                _money(row["margin"]),
                _pct(row["margin_pct"]),
            ]
        )
    _autosize(ws)


def _fill_category_sheet(ws: Worksheet, categories: list[dict]) -> None:
    _write_header(
        ws, ["Categoría", "Cantidad", "Ingreso", "Costo", "Margen $", "Margen %"]
    )
    for row in categories:
        ws.append(
            [
                row["name"],
                row["quantity"],
                _money(row["revenue"]),
                _money(row["cost"]),
                _money(row["margin"]),
                _pct(row["margin_pct"]),
            ]
        )
    _autosize(ws)


def _fill_rankings_sheet(ws: Worksheet, products: list[dict], top_n: int) -> None:
    by_revenue = products[:top_n]  # product_report is already revenue-sorted
    by_margin = sorted(products, key=lambda r: r["margin"], reverse=True)[:top_n]

    ws.append([f"Top {top_n} por ingreso"])
    ws["A1"].font = _HEADER_FONT
    ws.append(["Producto", "Ingreso"])
    for cell in ws[2]:
        cell.font = _HEADER_FONT
    for row in by_revenue:
        ws.append([row["name"], _money(row["revenue"])])

    ws.append([])
    margin_title_row = ws.max_row + 1
    ws.append([f"Top {top_n} por margen $"])
    ws.cell(row=margin_title_row, column=1).font = _HEADER_FONT
    header_row = ws.max_row + 1
    ws.append(["Producto", "Margen $"])
    for cell in ws[header_row]:
        cell.font = _HEADER_FONT
    for row in by_margin:
        ws.append([row["name"], _money(row["margin"])])

    _autosize(ws)


def _autosize(ws: Worksheet) -> None:
    """Widen each column to fit its longest value (capped)."""
    for column_cells in ws.columns:
        length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=0,
        )
        letter = column_cells[0].column_letter
        ws.column_dimensions[letter].width = min(max(length + 2, 10), 40)
