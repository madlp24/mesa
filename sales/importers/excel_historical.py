"""Importer for historical sales delivered as an Excel workbook.

One worksheet row becomes one :class:`CanonicalSale` carrying a single
:class:`CanonicalSaleItem`. The first row is treated as a header and columns
are matched by name, so column order does not matter. Rows missing any required
field are skipped and logged with their 1-based worksheet row number; the count
is exposed via :attr:`skipped_rows` for the caller's summary.
"""
import logging
from pathlib import Path

from openpyxl import load_workbook

from .base import BaseImporter
from .canonical import CanonicalSale
from .registry import register
from .rows import RowError, canonical_from_record

logger = logging.getLogger(__name__)


@register(".xlsx")
class ExcelHistoricalImporter(BaseImporter):
    """Parse a historical-sales Excel workbook into canonical sales."""

    def __init__(self) -> None:
        self.skipped_rows = 0

    def normalize(self, path: Path) -> list[CanonicalSale]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            rows = workbook.active.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                return []
            columns = [str(name).strip() if name is not None else "" for name in header]

            sales = []
            for row_number, row in enumerate(rows, start=2):
                record = dict(zip(columns, row))
                try:
                    sales.append(canonical_from_record(record))
                except RowError as exc:
                    self.skipped_rows += 1
                    logger.warning("Row %s skipped: %s", row_number, exc)
            return sales
        finally:
            workbook.close()
